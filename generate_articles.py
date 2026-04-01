#!/usr/bin/env python3
"""
MathSolver.cloud — Article Generator v4
Two-step generation: long text first, then structure into JSON
"""

import os, json, re, time, subprocess
from pathlib import Path
from datetime import datetime
import openpyxl

try:
    from openai import OpenAI
except ImportError:
    subprocess.run(["pip", "install", "openai", "openpyxl"], check=True)
    from openai import OpenAI

# ── CONFIG ───────────────────────────────────────────────────────────────────

OPENAI_API_KEY   = os.environ.get("OPENAI_API_KEY", "OPENAI_API_KEY")
GEN_MODEL        = "gpt-4o"
CHECK_MODEL      = "gpt-4o-mini"
ARTICLES_PER_RUN = int(os.environ.get("ARTICLES_PER_RUN", "3"))
MIN_SCORE        = 7
MAX_RETRIES      = 3

SITE_URL    = "https://mathsolver.cloud"
CWS_URL     = "https://chromewebstore.google.com/detail/math-solver/pieobakkfhafplomcoiohhpikcofoghb?utm_source=mathsolver.cloud&utm_medium=blog&utm_campaign=article"
CONTENT_PLAN   = Path("mathsolver_content_plan_FULL.xlsx")
OUTPUT_DIR     = Path("static/blog")
NEEDS_REVIEW   = Path("_needs_review")
PROGRESS_FILE  = Path(".generation_progress.json")

PILLAR_MAP = {
    "Algebra":                  "/algebra-solver/",
    "Geometry":                 "/geometry-solver/",
    "Calculus":                 "/calculus-solver/",
    "Trigonometry":             "/trigonometry-solver/",
    "Equation Solving":         "/equation-solver/",
    "Statistics & Probability": "/statistics-solver/",
    "Arithmetic & Fractions":   "/fractions-calculator/",
    "Matrix & Linear Algebra":  "/matrix-solver/",
    "Physics & Formulas":       "/physics-solver/",
    "Word Problems & Homework": "/math-word-problem-solver/",
}

# ── HELPERS ───────────────────────────────────────────────────────────────────

def load_progress():
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text())
    return {"published": [], "needs_review": []}

def save_progress(p):
    PROGRESS_FILE.write_text(json.dumps(p, indent=2))

def make_slug(keyword):
    slug = keyword.lower().strip()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'\s+', '-', slug)
    return re.sub(r'-+', '-', slug).strip('-')

def parse_json(raw):
    raw = raw.strip()
    # Remove markdown code blocks
    raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)
    raw = raw.strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # Fix invalid backslash escapes (LaTeX: \frac, \sqrt etc)
        raw = re.sub(r'(?<!\\)\\(?!["\\/bfnrtu])', r'\\\\', raw)
        return json.loads(raw)

def load_content_plan():
    wb = openpyxl.load_workbook(CONTENT_PLAN)
    ws = wb.active
    headers = [c.value for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
    articles = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            articles.append(dict(zip(headers, row)))
    return articles

# ── SCREENSHOT MAPPING ───────────────────────────────────────────────────────

SCREENSHOT_PREFIX = {
    "Algebra":                  "ALG",
    "Geometry":                 "GEO",
    "Calculus":                 "CALC",
    "Trigonometry":             "TRIG",
    "Equation Solving":         "EQ",
    "Statistics & Probability": "STAT",
    "Arithmetic & Fractions":   "FRAC",
    "Matrix & Linear Algebra":  "MAT",
    "Physics & Formulas":       "PHYS",
    "Word Problems & Homework": "WP",
}

def get_screenshot_html(cluster, example_num, screenshot_filename=None):
    if screenshot_filename:
        img_path = f"/mathsolver_solution_images/{screenshot_filename}"
    else:
        prefix = SCREENSHOT_PREFIX.get(cluster, "")
        if not prefix:
            return ""
        img_id = f"{prefix}-00{example_num}"
        img_path = f"/mathsolver_solution_images/{img_id}.png"
    return f'''
  <div class="ms-screenshot-wrap">
    <img src="{img_path}" alt="MathSolver solving example {example_num} — {cluster}" 
         width="400" height="300" loading="lazy"
         onerror="this.parentElement.style.display='none'">
    <p class="ms-screenshot-caption">MathSolver Chrome extension solving this problem step-by-step</p>
  </div>'''

# ── LOAD EXAMPLES FROM XLSX ──────────────────────────────────────────────────

EXAMPLES_FILE = Path("article_examples.xlsx")

def load_article_examples():
    """Load pre-generated examples from article_examples.xlsx"""
    if not EXAMPLES_FILE.exists():
        return {}
    
    wb = openpyxl.load_workbook(EXAMPLES_FILE)
    ws = wb.active
    examples = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        screenshot_id = row[0]  # e.g. STAT-A001
        filename      = row[1]  # e.g. STAT-A001.png
        slug          = row[2]  # e.g. mystatlab-homework-answers-statistics
        example_num   = row[5]  # 1 or 2
        problem       = row[6]  # the problem text
        
        if slug not in examples:
            examples[slug] = {}
        examples[slug][example_num] = {
            "screenshot_id": screenshot_id,
            "filename": filename,
            "problem": problem
        }
    
    return examples

# Load examples once at startup
ARTICLE_EXAMPLES = load_article_examples()

# ── STEP 1: WRITE LONG ARTICLE (plain text) ───────────────────────────────────

def prompt_write_article(keyword, title, cluster, related_links, pillar_url):
    return f"""You are an expert math tutor and SEO writer for MathSolver.cloud.

Write a comprehensive, detailed article about "{keyword}".

REQUIREMENTS:
- Total length: MINIMUM 1,500 words (aim for 1,800-2,000 words)
- Tone: Friendly and clear, like a helpful tutor explaining to a student
- Start the article directly with the topic — NO generic intros like "In today's world..."
- The keyword "{keyword}" must appear in the very first sentence
- IMPORTANT: Write ALL formulas in plain text only. NO LaTeX, NO backslashes, NO \frac, NO \sum, NO \sqrt. Write fractions as "a/b", square roots as "sqrt(x)", summation as "sum of...".

WRITE THESE SECTIONS IN ORDER:

## Introduction (3 full paragraphs, ~200 words)
Explain what {keyword} is, why students struggle with it, and what they'll learn. Keyword in first sentence.

## Key Formula or Definition (1 paragraph)
State the main formula or definition clearly.

## Step-by-Step Guide (4 steps, each step = 1 full paragraph of 5-6 sentences)
Step 1: [descriptive title]
Step 2: [descriptive title]
Step 3: [descriptive title]
Step 4: [descriptive title]

## Worked Example 1 (full detailed solution, ~150 words)
Present a concrete math problem then solve it step by step showing all work.

## Worked Example 2 (full detailed solution, ~150 words)
A slightly harder problem, solved step by step.

## Common Mistakes to Avoid (2 full paragraphs, ~150 words)
Describe the most common errors students make and how to avoid them.

## Real-World Applications (2 full paragraphs, ~150 words)
Where is {keyword} used in real life? Give specific examples.

## FAQ (5 questions with detailed answers, each answer 3-4 sentences)
Q1: A question containing "{keyword}"
Q2: A common student question about this topic
Q3: How can AI help with {keyword}? (mention MathSolver Chrome extension — take a screenshot, get instant step-by-step solution)
Q4: Another relevant question
Q5: A practical question

RELATED TOPICS TO MENTION:
{related_links if related_links else "General math topics"}

PILLAR PAGE: {pillar_url}
Mention this page once naturally as "our complete {cluster} guide".

Write the full article now. Use clear headers. Be thorough and detailed."""

# ── STEP 2: STRUCTURE INTO JSON ───────────────────────────────────────────────

def prompt_structure_json(article_text, keyword):
    return f"""Convert this article into a structured JSON object. Extract and organize the content exactly as written — do NOT summarize or shorten anything.

ARTICLE TEXT:
{article_text[:8000]}

Return ONLY this JSON structure with NO markdown, NO backticks:
{{
  "meta_title": "under 60 chars, contains '{keyword}', ends with year 2026",
  "meta_description": "under 160 chars, contains keyword, mentions MathSolver AI",
  "h1": "the main article headline containing '{keyword}'",
  "intro": "the full introduction text with paragraphs separated by \\n\\n (double newline). Keep all 3 paragraphs separated.",
  "formula": "ONE SHORT LINE only — the main formula, e.g. Y = a + bX or Mean = sum / n. Max 50 chars. No explanations.",
  "formula_label": "short label like 'Standard Formula'",
  "steps": [
    {{"title": "step title", "content": "the FULL step explanation — do not shorten, include all sentences"}},
    {{"title": "step title", "content": "the FULL step explanation"}},
    {{"title": "step title", "content": "the FULL step explanation"}},
    {{"title": "step title", "content": "the FULL step explanation"}}
  ],
  "example1_problem": "the problem statement from Example 1",
  "example1_solution": "the COMPLETE solution with each step on its own line, starting with Step 1:, Step 2:, etc. Use \\n between steps.",
  "example2_problem": "the problem statement from Example 2",
  "example2_solution": "the COMPLETE solution with each step on its own line, starting with Step 1:, Step 2:, etc. Use \\n between steps.",
  "common_mistakes": "the FULL common mistakes text with paragraphs separated by \\n\\n",
  "real_world": "the FULL real-world applications text with paragraphs separated by \\n\\n",
  "faq": [
    {{"q": "question 1", "a": "full answer"}},
    {{"q": "question 2", "a": "full answer"}},
    {{"q": "question 3 about AI/MathSolver", "a": "full answer mentioning MathSolver"}},
    {{"q": "question 4", "a": "full answer"}},
    {{"q": "question 5", "a": "full answer"}}
  ],
  "lsi_keywords": ["related term 1", "related term 2", "related term 3", "related term 4", "related term 5"]
}}"""

# ── STEP 3: VALIDATE (Python-based, no GPT) ──────────────────────────────────

def validate_article(article_json, keyword):
    """Validate article quality using Python checks — fast and reliable."""
    scores = {}
    failed = []

    # 1. Has h1
    scores["has_h1"] = 1 if article_json.get("h1","").strip() else 0

    # 2. Has intro with keyword
    intro = article_json.get("intro","")
    intro_words = intro.split()[:150]
    scores["keyword_in_intro"] = 1 if keyword.lower() in " ".join(intro_words).lower() else 0

    # 3. Has steps (at least 3)
    steps = article_json.get("steps", [])
    scores["has_steps"] = 1 if len(steps) >= 3 else 0

    # 4. Has examples
    ex1 = article_json.get("example1_problem","") or article_json.get("example1_solution","")
    ex2 = article_json.get("example2_problem","") or article_json.get("example2_solution","")
    scores["has_examples"] = 1 if ex1 and ex2 else 0

    # 5. Has FAQ (at least 3)
    faq = article_json.get("faq", [])
    scores["has_faq"] = 1 if len(faq) >= 3 else 0

    # 6. Sufficient word count (500+ words)
    all_text = " ".join([
        intro,
        " ".join(s.get("content","") for s in steps),
        article_json.get("example1_solution",""),
        article_json.get("example2_solution",""),
        article_json.get("common_mistakes",""),
        article_json.get("real_world",""),
        " ".join(f.get("a","") for f in faq)
    ])
    word_count = len(all_text.split())
    scores["sufficient_length"] = 1 if word_count >= 500 else 0

    # 7. MathSolver mentioned in FAQ
    faq_text = " ".join(f.get("a","") for f in faq).lower()
    scores["mentions_mathsolver"] = 1 if "mathsolver" in faq_text else 0

    total = sum(scores.values())
    failed = [k for k,v in scores.items() if v == 0]

    return total, failed, word_count

# ── BUILD HTML ────────────────────────────────────────────────────────────────

def build_html(data, keyword, slug, cluster, url):
    pillar_url  = PILLAR_MAP.get(cluster, "/")
    pub_date    = datetime.now().strftime("%B %Y")
    year        = datetime.now().year

    # Fallbacks
    h1    = data.get("h1") or keyword.title() + " — Step-by-Step Guide"
    title = data.get("meta_title") or f"{keyword.title()} Guide ({year})"
    desc  = data.get("meta_description") or f"Learn {keyword} step by step. Free AI math solver."
    intro = data.get("intro") or f"Learn how to solve {keyword} with this complete guide."
    formula       = data.get("formula") or "See step-by-step solution below"
    formula_label = data.get("formula_label") or "Key Concept"

    # Steps HTML
    steps_html = ""
    for i, step in enumerate(data.get("steps", []), 1):
        steps_html += f"""
  <div class="ms-step">
    <div class="ms-step-num">{i}</div>
    <div class="ms-step-body">
      <h3>{step.get('title','')}</h3>
      <p>{step.get('content','')}</p>
    </div>
  </div>"""

    # FAQ HTML
    faq_html = ""
    for i, item in enumerate(data.get("faq", [])):
        open_attr = "open" if i == 0 else ""
        faq_html += f"""
  <details class="ms-faq-item" {open_attr}>
    <summary class="ms-faq-q">❓ {item.get('q','')}</summary>
    <div class="ms-faq-a">{item.get('a','')}</div>
  </details>"""

    # Schema
    faq_schema = ""
    for item in data.get("faq", []):
        q = item.get('q','').replace('"',"'")
        a = item.get('a','').replace('"',"'")
        faq_schema += f'{{"@type":"Question","name":"{q}","acceptedAnswer":{{"@type":"Answer","text":"{a}"}}}},'

    def format_solution(sol):
        if not sol:
            return ""
        # Split on Step N: pattern
        steps = re.split(r'(?=Step \d+:)', sol)
        parts = [s.strip() for s in steps if s.strip()]
        if len(parts) > 1:
            return "".join(f'<div class="sol-step">{p}</div>' for p in parts)
        return sol.replace("\n", "<br>")
    ex1_sol = format_solution(data.get("example1_solution",""))
    ex2_sol = format_solution(data.get("example2_solution",""))
    # Get screenshot filenames from article_examples.xlsx
    art_examples = ARTICLE_EXAMPLES.get(slug, {})
    sc1_file = art_examples.get(1, {}).get("filename")
    sc2_file = art_examples.get(2, {}).get("filename")
    screenshot1 = get_screenshot_html(cluster, 1, sc1_file)
    screenshot2 = get_screenshot_html(cluster, 2, sc2_file)
    mistakes_raw = data.get("common_mistakes","")
    mistakes = "</p><p>".join([p.strip() for p in re.split(r'\n\n+', mistakes_raw) if p.strip()]) if mistakes_raw else ""
    real_world_raw = data.get("real_world","")
    real_world = "</p><p>".join([p.strip() for p in re.split(r'\n\n+', real_world_raw) if p.strip()]) if real_world_raw else ""
    
    # Split intro into paragraphs properly
    intro_parts = [p.strip() for p in re.split(r'\n\n+', intro) if p.strip()]
    intro_html = "</p><p>".join(intro_parts)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title}</title>
  <meta name="description" content="{desc}">
  <link rel="canonical" href="{SITE_URL}/blog/{slug}/">
  <meta property="og:title" content="{title}">
  <meta property="og:description" content="{desc}">
  <meta property="og:url" content="{SITE_URL}/blog/{slug}/">
  <meta property="og:type" content="article">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link href="https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,700;12..96,800;12..96,900&family=DM+Sans:opsz,wght@0,9..40,400;0,9..40,500&display=swap" rel="stylesheet">
  <script async src="https://www.googletagmanager.com/gtag/js?id=G-5DLHQQDXLW"></script>
  <script>window.dataLayer=window.dataLayer||[];function gtag(){{dataLayer.push(arguments)}}gtag("js",new Date());gtag("config","G-5DLHQQDXLW");</script>
  <script type="application/ld+json">{{
    "@context":"https://schema.org","@graph":[
      {{"@type":"Article","headline":"{h1.replace('"',"'")}","description":"{desc.replace('"',"'")}","datePublished":"{datetime.now().strftime('%Y-%m-%d')}","author":{{"@type":"Organization","name":"MathSolver Team"}},"publisher":{{"@type":"Organization","name":"MathSolver","url":"{SITE_URL}"}}}},
      {{"@type":"FAQPage","mainEntity":[{faq_schema.rstrip(',')}]}},
      {{"@type":"BreadcrumbList","itemListElement":[{{"@type":"ListItem","position":1,"name":"Home","item":"{SITE_URL}"}},{{"@type":"ListItem","position":2,"name":"{cluster}","item":"{SITE_URL}{pillar_url}"}},{{"@type":"ListItem","position":3,"name":"{h1.replace('"',"'")}","item":"{SITE_URL}/blog/{slug}/"}}]}}
    ]
  }}</script>
  <style>
    *,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
    :root{{--blue:#2B7FE8;--blue-dark:#1a5fc4;--yellow:#FFB800;--dark:#060D1F;--dark-2:#0D1B35;--dark-3:#112248;--white:#fff;--gray:#8a97b0;--text:#d4dff5;--border:rgba(43,127,232,.15);--r:16px}}
    html{{scroll-behavior:smooth}}
    body{{font-family:'DM Sans',sans-serif;background:var(--dark);color:var(--text);line-height:1.75;overflow-x:hidden;text-align:left}}
    h1,h2,h3,h4{{font-family:'Bricolage Grotesque',sans-serif;line-height:1.2}}
    a{{color:var(--blue);text-decoration:none}}
    a:hover{{text-decoration:underline}}
    header{{position:fixed;top:0;left:0;right:0;z-index:100;height:72px;padding:0 40px;display:flex;align-items:center;justify-content:space-between;background:rgba(6,13,31,.92);backdrop-filter:blur(20px);border-bottom:1px solid rgba(43,127,232,.1)}}
    .logo{{font-family:'Bricolage Grotesque',sans-serif;font-size:1.3rem;font-weight:800;color:#fff;display:flex;align-items:center;gap:10px}}
    .logo-icon{{width:34px;height:34px;background:linear-gradient(135deg,var(--blue),var(--yellow));border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:18px}}
    nav{{display:flex;align-items:center;gap:4px}}
    nav a{{color:var(--gray);font-size:.875rem;padding:8px 14px;border-radius:8px;transition:.2s;white-space:nowrap}}
    nav a:hover{{color:#fff;background:rgba(255,255,255,.06)}}
    .btn-nav{{background:var(--yellow)!important;color:var(--dark)!important;font-weight:700!important;border-radius:50px!important;margin-left:8px}}
    .hamburger{{display:none;flex-direction:column;gap:5px;cursor:pointer;background:none;border:none;padding:4px}}
    .hamburger span{{width:24px;height:2px;background:#fff;border-radius:2px;display:block}}
    .wrap{{max-width:820px;margin:0 auto;padding:100px 40px 80px}}
    .ms-breadcrumb{{font-size:13px;color:#666;margin-bottom:24px;padding-bottom:12px;border-bottom:1px solid var(--border)}}
    .ms-breadcrumb a{{color:var(--blue)}}
    .ms-breadcrumb span{{margin:0 6px;color:#444}}
    .ms-hero{{background:linear-gradient(135deg,#1A1A2E 0%,#2B7FE8 100%);border-radius:20px;padding:48px 40px;margin-bottom:32px;color:#fff}}
    .ms-hero h1{{font-size:clamp(1.6rem,3vw,2.3rem);font-weight:900;margin:0 0 16px;color:#fff}}
    .ms-meta{{font-size:13px;opacity:.75;display:flex;gap:20px;flex-wrap:wrap;margin-top:16px}}
    .ms-toc{{background:rgba(43,127,232,.08);border-left:4px solid var(--blue);border-radius:0 12px 12px 0;padding:24px 28px;margin-bottom:32px}}
    .ms-toc h3{{font-size:14px;font-weight:700;color:#fff;margin:0 0 12px;text-transform:uppercase;letter-spacing:.5px}}
    .ms-toc ol{{margin:0;padding-left:20px}}
    .ms-toc li{{margin-bottom:6px}}
    .ms-toc a{{color:var(--blue);font-size:14px;font-weight:500}}
    .ms-intro{{font-size:17px;line-height:1.8;margin-bottom:32px;text-align:left}}
    .ms-intro p{{margin-bottom:16px;text-align:left}}
    .ms-formula{{background:var(--dark-3);border-radius:12px;padding:28px 32px;margin:28px 0;text-align:center;border:1px solid var(--border)}}
    .ms-formula .formula{{font-size:clamp(1.1rem,2vw,1.6rem);color:#fff;font-family:'Courier New',monospace;font-weight:700}}
    .ms-formula .label{{font-size:12px;color:var(--yellow);text-transform:uppercase;letter-spacing:1px;margin-top:10px}}
    .ms-steps{{margin:32px 0}}
    .ms-steps h2{{font-size:1.4rem;font-weight:800;color:#fff;margin-bottom:20px}}
    .ms-step{{display:flex;gap:20px;margin-bottom:20px;padding:24px;background:var(--dark-2);border:1px solid var(--border);border-radius:12px}}
    .ms-step-num{{width:44px;height:44px;min-width:44px;background:var(--blue);color:#fff;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:18px;font-family:'Bricolage Grotesque',sans-serif}}
    .ms-step-body h3{{font-size:16px;font-weight:700;color:#fff;margin:0 0 8px;text-align:left}}
    .ms-step-body p{{font-size:15px;line-height:1.7;color:var(--gray);margin:0;text-align:left}}
    .ms-cta{{background:linear-gradient(135deg,var(--blue) 0%,var(--blue-dark) 100%);border-radius:16px;padding:40px;text-align:center;margin:36px 0;color:#fff}}
    .ms-cta h3{{font-size:1.4rem;font-weight:800;margin:0 0 10px;color:#fff}}
    .ms-cta p{{opacity:.85;margin:0 0 20px;font-size:16px}}
    .ms-cta-btn{{display:inline-block;background:var(--yellow);color:var(--dark);font-weight:800;font-size:16px;padding:14px 36px;border-radius:50px}}
    .ms-cta-btn:hover{{transform:scale(1.04);color:var(--dark);text-decoration:none}}
    .ms-examples{{margin:32px 0}}
    .ms-examples h2{{font-size:1.4rem;font-weight:800;color:#fff;margin-bottom:20px}}
    .ms-example{{background:var(--dark-2);border:1px solid var(--border);border-radius:12px;padding:28px;margin-bottom:20px}}
    .ms-example h3{{font-size:1.05rem;font-weight:700;color:#fff;margin-bottom:12px}}
    .ms-example .problem{{font-size:1rem;font-weight:600;color:var(--yellow);margin-bottom:14px}}
    .ms-example .solution{{font-size:.95rem;line-height:1.85;color:var(--text)}}
    .sol-step{{padding:8px 0;border-bottom:1px solid rgba(255,255,255,.06);margin-bottom:4px}}
    .sol-step:last-child{{border-bottom:none}}
    .ms-screenshot-wrap{{margin:20px 0;text-align:center}}
    .ms-screenshot-wrap img{{max-width:400px;width:100%;height:auto;border-radius:12px;box-shadow:0 4px 24px rgba(0,0,0,.3);border:1px solid var(--border)}}
    .ms-screenshot-caption{{font-size:13px;color:var(--gray);margin-top:8px;font-style:italic}}
    .ms-extra{{margin:32px 0}}
    .ms-extra h2{{font-size:1.4rem;font-weight:800;color:#fff;margin-bottom:16px}}
    .ms-extra p{{font-size:15px;line-height:1.8;color:var(--gray);margin-bottom:12px;text-align:left}}
    .ms-faq{{margin:36px 0}}
    .ms-faq h2{{font-size:1.4rem;font-weight:800;color:#fff;margin-bottom:20px}}
    .ms-faq-item{{border:1px solid var(--border);border-radius:12px;margin-bottom:10px;overflow:hidden}}
    .ms-faq-item summary{{background:var(--dark-2);padding:18px 22px;font-weight:700;font-size:15px;color:#fff;cursor:pointer;list-style:none;display:flex;align-items:center;gap:8px;text-align:left}}
    .ms-faq-item summary::-webkit-details-marker{{display:none}}
    .ms-faq-a{{padding:18px 22px;font-size:15px;line-height:1.7;color:var(--gray);border-top:1px solid var(--border);text-align:left}}
    .ms-rating{{text-align:center;padding:28px;background:rgba(43,127,232,.07);border-radius:12px;margin:32px 0;border:1px solid var(--border)}}
    .ms-stars{{font-size:28px}}
    .ms-related{{background:var(--dark-2);border-radius:12px;padding:28px;margin:32px 0;border:1px solid var(--border)}}
    .ms-related h3{{font-size:16px;font-weight:700;color:#fff;margin:0 0 16px;text-align:left}}
    .ms-related ul{{list-style:none;display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px}}
    .ms-related a{{display:block;padding:10px 14px;background:var(--dark-3);border:1px solid var(--border);border-radius:8px;font-size:14px;font-weight:500;color:var(--blue);transition:.2s}}
    .ms-related a:hover{{border-color:var(--blue);text-decoration:none}}
    footer{{border-top:1px solid rgba(255,255,255,.06);padding:48px 0 32px;margin-top:40px}}
    .footer-inner{{max-width:1160px;margin:0 auto;padding:0 40px;display:flex;flex-direction:column;align-items:center;gap:24px}}
    .footer-logo{{font-family:'Bricolage Grotesque',sans-serif;font-weight:700;font-size:1.2rem;color:#fff}}
    .footer-nav{{display:flex;flex-wrap:wrap;justify-content:center;gap:4px}}
    .footer-nav a{{color:var(--gray);font-size:.875rem;padding:6px 14px;border-radius:8px;transition:.2s}}
    .footer-nav a:hover{{color:#fff}}
    .footer-copy{{color:rgba(255,255,255,.25);font-size:.8rem;text-align:center;line-height:1.8}}
    @media(max-width:860px){{
      header{{padding:0 20px}}
      nav{{display:none}}
      nav.open{{display:flex;flex-direction:column;position:fixed;top:72px;left:0;right:0;background:rgba(6,13,31,.98);padding:20px;border-bottom:1px solid var(--border)}}
      nav.open a{{padding:12px 16px;font-size:1rem}}
      nav.open .btn-nav{{margin-left:0;margin-top:8px}}
      .hamburger{{display:flex}}
      .wrap{{padding:90px 20px 60px}}
      .ms-hero{{padding:32px 24px}}
      .ms-step{{flex-direction:column;gap:12px}}
    }}
  </style>
</head>
<body>
  <header>
    <a href="/" class="logo"><span class="logo-icon">&#8721;</span>MathSolver</a>
    <button class="hamburger" id="ham" aria-label="Menu"><span></span><span></span><span></span></button>
    <nav id="nav">
      <a href="/">Home</a>
      <a href="{CWS_URL}" target="_blank" rel="noopener" class="btn-nav">⚡ Add to Chrome — Free</a>
    </nav>
  </header>

  <div class="wrap">

    <nav class="ms-breadcrumb" aria-label="Breadcrumb">
      <a href="/">Home</a><span>›</span>
      <a href="{SITE_URL}{pillar_url}">{cluster}</a><span>›</span>
      <span>{h1}</span>
    </nav>

    <div class="ms-hero">
      <h1>{h1}</h1>
      <div class="ms-meta">
        <span>📅 Updated {pub_date}</span>
        <span>⏱ 8 min read</span>
        <span>🎓 All levels</span>
        <span>✍️ By MathSolver Team</span>
      </div>
    </div>

    <div class="ms-toc">
      <h3>📋 In this guide</h3>
      <ol>
        <li><a href="#intro">What is {keyword.title()}?</a></li>
        <li><a href="#formula">Key Formula</a></li>
        <li><a href="#steps">Step-by-Step Guide</a></li>
        <li><a href="#examples">Worked Examples</a></li>
        <li><a href="#mistakes">Common Mistakes</a></li>
        <li><a href="#real-world">Real-World Uses</a></li>
        <li><a href="#solver">Try AI Solver</a></li>
        <li><a href="#faq">FAQ</a></li>
      </ol>
    </div>

    <div class="ms-intro" id="intro">
      <p>{intro_html}</p>
    </div>

    <div class="ms-formula" id="formula">
      <div class="formula">{formula}</div>
      <div class="label">{formula_label}</div>
    </div>

    <div class="ms-steps" id="steps">
      <h2>Step-by-Step: How to Solve {keyword.title()}</h2>
      {steps_html}
    </div>

    <div class="ms-cta" id="solver">
      <h3>🤖 Stuck on a math problem?</h3>
      <p>Take a screenshot and let our AI solve it step-by-step in seconds</p>
      <a href="{CWS_URL}" class="ms-cta-btn" target="_blank" rel="noopener">⚡ Try MathSolver Free →</a>
    </div>

    <div class="ms-examples" id="examples">
      <h2>Worked Examples</h2>
      <div class="ms-example">
        <h3>Example 1</h3>
        <div class="problem">Problem: {data.get('example1_problem','')}</div>
        <div class="solution">{ex1_sol}</div>
        {screenshot1}
      </div>
      <div class="ms-example">
        <h3>Example 2</h3>
        <div class="problem">Problem: {data.get('example2_problem','')}</div>
        <div class="solution">{ex2_sol}</div>
        {screenshot2}
      </div>
    </div>

    <div class="ms-extra" id="mistakes">
      <h2>Common Mistakes to Avoid</h2>
      <p>{mistakes}</p>
    </div>

    <div class="ms-extra" id="real-world">
      <h2>Real-World Applications</h2>
      <p>{real_world}</p>
    </div>

    <div class="ms-faq" id="faq">
      <h2>Frequently Asked Questions</h2>
      {faq_html}
    </div>

    <div class="ms-rating">
      <p>Was this guide helpful?</p>
      <div class="ms-stars">⭐⭐⭐⭐⭐</div>
      <p style="margin-top:10px;font-size:13px;color:var(--gray)">4.8/5 based on 127 ratings</p>
    </div>

    <div class="ms-related" id="related">
      <h3>📚 Related Topics</h3>
      <ul>
        <li><a href="{SITE_URL}{pillar_url}">{cluster} Solver — Complete Guide</a></li>
      </ul>
    </div>

    <div class="ms-cta">
      <h3>🚀 Solve any math problem instantly</h3>
      <p>2,000+ students use MathSolver every day — join them for free</p>
      <a href="{CWS_URL}" class="ms-cta-btn" target="_blank" rel="noopener">📥 Add to Chrome — It's Free</a>
    </div>

  </div>

  <footer>
    <div class="footer-inner">
      <a href="/" class="footer-logo">MathSolver</a>
      <nav class="footer-nav">
        <a href="/">Home</a>
        <a href="/privacy-policy/">Privacy notice</a>
        <a href="/refund-policy/">Refund policy</a>
        <a href="/price/">Price</a>
        <a href="/terms-of-service/">Terms of Service</a>
      </nav>
      <div class="footer-copy">Copyright &copy; 2024 &ldquo;MARGOAPPS&rdquo; LLC<br>Armenia, Yervand Kochar Street, 8 &mdash; Yerevan, 0070</div>
    </div>
  </footer>

  <!-- Paddle not needed on blog pages -->
  <script>
    const h=document.getElementById('ham'),n=document.getElementById('nav');
    if(h)h.addEventListener('click',()=>n.classList.toggle('open'));
    document.addEventListener('click',e=>{{if(h&&!h.contains(e.target)&&!n.contains(e.target))n.classList.remove('open')}});
  </script>
</body>
</html>"""

# ── GENERATE ARTICLE ──────────────────────────────────────────────────────────

def generate_article(client, article, related):
    keyword = article.get("Primary Keyword","")
    cluster = article.get("Cluster","")
    title   = article.get("Article Title","")
    url     = article.get("URL","")
    pillar  = PILLAR_MAP.get(cluster,"/")

    related_links = "\n".join([f'- {r["keyword"]}' for r in related[:4]])

    article_data = {}
    score = 0

    for attempt in range(1, MAX_RETRIES+1):
        print(f"\nAttempt {attempt}/{MAX_RETRIES}...")

        try:
            # Step 1: Write long article
            print("  [1/3] Writing article with GPT-4o...")
            # Get pre-defined examples for this article
            art_examples = ARTICLE_EXAMPLES.get(slug, {})
            ex1_data = art_examples.get(1, {})
            ex2_data = art_examples.get(2, {})
            ex1_problem = ex1_data.get("problem")
            ex2_problem = ex2_data.get("problem")
            if ex1_problem:
                print(f"  → Using pre-defined examples: {ex1_data.get('screenshot_id')} + {ex2_data.get('screenshot_id')}")
            r1 = client.chat.completions.create(
                model=GEN_MODEL,
                messages=[{"role":"user","content": prompt_write_article(keyword, title, cluster, related_links, f"{SITE_URL}{pillar}", ex1_problem, ex2_problem)}],
                temperature=0.7,
                max_tokens=6000,
            )
            article_text = r1.choices[0].message.content.strip()
            word_count = len(article_text.split())
            print(f"  → Article text: {word_count} words")

            # Step 2: Structure into JSON
            print("  [2/3] Structuring into JSON...")
            r2 = client.chat.completions.create(
                model=GEN_MODEL,
                messages=[{"role":"user","content": prompt_structure_json(article_text, keyword)}],
                temperature=0.2,
                max_tokens=6000,
            )
            article_data = parse_json(r2.choices[0].message.content)

            # Step 3: Validate (Python-based)
            print("  [3/3] Quality check...")
            score, failed, wc = validate_article(article_data, keyword)

            print(f"  Score: {score}/7 | Words: {wc}")
            if failed:
                print(f"  Failed: {', '.join(failed)}")

            if score >= MIN_SCORE:
                print(f"  ✓ Quality check passed! ({score}/7)")
                return article_data, score

        except Exception as e:
            print(f"  ✗ Error: {e}")

        time.sleep(2)

    return article_data, score

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print(f"MathSolver Article Generator v4 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

    if OPENAI_API_KEY == "OPENAI_API_KEY":
        print("❌ OPENAI_API_KEY not set!")
        return

    client = OpenAI(api_key=OPENAI_API_KEY)

    articles = load_content_plan()
    progress = load_progress()
    published_urls = {a["url"] for a in progress["published"]}
    review_urls = {a["url"] for a in progress.get("needs_review",[])}

    sub_articles = [
        a for a in articles
        if a.get("Type") == "SUB-ARTICLE"
        and a.get("URL") not in published_urls
        and a.get("URL") not in review_urls
    ]

    # Phase 1 priority
    try:
        wb = openpyxl.load_workbook(CONTENT_PLAN)
        ws1 = wb['🚀 Quick Wins (Phase 1)']
        phase1_keys = {row[3] for row in ws1.iter_rows(min_row=4, values_only=True) if row[3]}
        priority = [a for a in sub_articles if a.get("Primary Keyword") in phase1_keys]
        rest = [a for a in sub_articles if a.get("Primary Keyword") not in phase1_keys]
        queue = (priority + rest)[:ARTICLES_PER_RUN]
    except:
        queue = sub_articles[:ARTICLES_PER_RUN]

    if not queue:
        print("✅ All articles published!")
        return

    print(f"📋 Queue: {len(queue)} articles | Published: {len(published_urls)} | Remaining: {len(sub_articles)}\n")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    NEEDS_REVIEW.mkdir(exist_ok=True)
    published_count = 0

    for article in queue:
        keyword = article.get("Primary Keyword","")
        cluster = article.get("Cluster","")
        url     = article.get("URL","")
        slug    = make_slug(keyword)
        if not url:
            url = f"/blog/{slug}"

        related = [p for p in progress["published"] if p.get("cluster")==cluster][:4]

        print(f"\n{'='*60}")
        print(f"Generating: {keyword}")
        print(f"Cluster: {cluster}")
        print(f"{'='*60}")

        article_data, score = generate_article(client, article, related)

        html = build_html(article_data, keyword, slug, cluster, url)

        if score >= MIN_SCORE:
            out_dir = OUTPUT_DIR / slug
            out_dir.mkdir(parents=True, exist_ok=True)
            (out_dir / "index.html").write_text(html, encoding="utf-8")
            progress["published"].append({
                "url": url, "slug": slug, "cluster": cluster,
                "keyword": keyword, "score": score,
                "published_at": datetime.now().isoformat()
            })
            save_progress(progress)
            published_count += 1
            print(f"\n  ✅ Published: /blog/{slug}/")
        else:
            (NEEDS_REVIEW / f"{slug}.html").write_text(html, encoding="utf-8")
            progress.setdefault("needs_review",[]).append({
                "url": url, "slug": slug, "score": score,
                "published_at": datetime.now().isoformat()
            })
            save_progress(progress)
            print(f"\n  ⚠ Needs review (score {score}/7): {slug}")

        time.sleep(2)

    print(f"\n{'='*60}")
    print(f"✅ Done! Published: {published_count}/{len(queue)}")
    print(f"{'='*60}\n")

    # Git operations handled by GitHub Actions workflow
    print(f"✅ Generation complete. GitHub Actions will commit and push.")

if __name__ == "__main__":
    main()
