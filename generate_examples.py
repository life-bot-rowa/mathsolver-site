#!/usr/bin/env python3
"""
MathSolver.cloud — Example Problem Generator
Generates 2 math problems per article → save to Excel → you solve in MathSolver → screenshot
"""

import os, json, re, time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from openai import OpenAI
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openai", "openpyxl"], check=True)
    from openai import OpenAI

# ── CONFIG ────────────────────────────────────────────────────────────────────

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "OPENAI_API_KEY")
MODEL          = "gpt-4o-mini"
CONTENT_PLAN   = Path("mathsolver_content_plan_FULL.xlsx")
OUTPUT_FILE    = Path("article_examples.xlsx")
PROGRESS_FILE  = Path("examples_progress.json")

CLUSTER_PREFIX = {
    "Algebra":                  "ALG",
    "Geometry":                 "GEO",
    "Calculus":                 "CALC",
    "Trigonometry":             "TRIG",
    "Equation Solving":         "EQ",
    "Statistics & Probability": "STAT",
    "Arithmetic & Fractions":   "FRAC",
    "Matrix & Linear Algebra":  "MAT",
    "Physics & Formulas":       "PHYS",
    "Word Problems & Homework": "WP",
}

# ── HELPERS ───────────────────────────────────────────────────────────────────

def load_articles():
    wb = openpyxl.load_workbook(CONTENT_PLAN)
    ws = wb.active
    headers = [c.value for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=3, values_only=True) if row[0] == "SUB-ARTICLE"]

def load_progress():
    if PROGRESS_FILE.exists():
        return json.loads(PROGRESS_FILE.read_text())
    return {"done": [], "counters": {}}

def save_progress(p):
    PROGRESS_FILE.write_text(json.dumps(p, indent=2))

def make_slug(keyword):
    slug = keyword.lower().strip()
    slug = re.sub(r'[^a-z0-9\s-]', '', slug)
    slug = re.sub(r'\s+', '-', slug)
    return re.sub(r'-+', '-', slug).strip('-')

# ── GENERATE PROBLEMS ─────────────────────────────────────────────────────────

def generate_problems(client, keyword, cluster):
    prompt = f"""Create exactly 2 math problems for the topic "{keyword}" ({cluster}).

Rules:
- Use specific real numbers (not abstract variables like a, b, c)
- Example 1 = easier, Example 2 = harder  
- Plain text only: use / for division, sqrt() for roots, ^ for exponents
- NO LaTeX, NO backslashes

Return ONLY valid JSON:
{{
  "problem1": "full problem statement with numbers",
  "problem2": "full harder problem statement with numbers"
}}"""

    try:
        r = client.chat.completions.create(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
            max_tokens=300,
        )
        raw = r.choices[0].message.content.strip()
        raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)
        return json.loads(raw)
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return None

# ── SAVE EXCEL ────────────────────────────────────────────────────────────────

def save_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Examples for Screenshots"

    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill("solid", fgColor="1A1A2E")
    b_fill = PatternFill("solid", fgColor="E8F1FD")
    wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    brd    = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Screenshot ID", "Screenshot Filename", "Article Slug", 
               "Cluster", "Keyword", "Example #", "Problem to solve in MathSolver", "Status"]
    widths  = [14, 22, 38, 22, 32, 10, 65, 12]

    ws.append(headers)
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = center
        cell.border = brd
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows, 2):
        fill = b_fill if row["example_num"] == 1 else PatternFill("solid", fgColor="FFFFFF")
        vals = [row["id"], row["filename"], row["slug"], row["cluster"],
                row["keyword"], row["example_num"], row["problem"], "📸 TODO"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(i, col, value=val)
            cell.fill = fill
            cell.border = brd
            cell.alignment = wrap
            cell.font = Font(size=10)

    # Instructions sheet
    ws2 = wb.create_sheet("How To")
    ws2.column_dimensions["A"].width = 70
    ws2["A1"] = "How to prepare screenshots"
    ws2["A1"].font = Font(bold=True, size=14)
    for i, line in enumerate([
        "", "WORKFLOW:",
        "1. Open MathSolver Chrome extension",
        "2. Type the problem from column 'Problem to solve in MathSolver'",
        "3. MathSolver shows step-by-step solution",
        "4. Screenshot the solution panel",
        "5. Name the file exactly as shown in 'Screenshot Filename' column",
        "6. Put all screenshots in: static/mathsolver_solution_images/",
        "7. git add . && git commit -m 'add screenshots' && git push",
        "8. Articles will automatically show the correct screenshot",
        "", "NOTE: Screenshot filename must match exactly (case sensitive)",
    ], 2):
        ws2.cell(i, 1, line)

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved to {OUTPUT_FILE} — {len(rows)} rows")

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print("MathSolver — Example Problem Generator")
    print(f"{'='*60}\n")

    if OPENAI_API_KEY == "OPENAI_API_KEY":
        print("❌ Set OPENAI_API_KEY")
        return

    client   = OpenAI(api_key=OPENAI_API_KEY)
    articles = load_articles()
    progress = load_progress()
    done     = set(progress["done"])
    counters = progress.get("counters", {})

    # Load existing rows
    existing = []
    if OUTPUT_FILE.exists():
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                existing.append({"id": row[0], "filename": row[1], "slug": row[2],
                                  "cluster": row[3], "keyword": row[4],
                                  "example_num": row[5], "problem": row[6]})

    pending = [a for a in articles if make_slug(a.get("Primary Keyword","")) not in done]
    print(f"Total: {len(articles)} | Done: {len(done)} | Pending: {len(pending)}")

    new_rows = []

    for i, article in enumerate(pending, 1):
        keyword = article.get("Primary Keyword", "")
        cluster = article.get("Cluster", "")
        slug    = make_slug(keyword)
        prefix  = CLUSTER_PREFIX.get(cluster, "GEN")

        print(f"\n[{i}/{len(pending)}] {keyword}")

        problems = generate_problems(client, keyword, cluster)
        if not problems:
            continue

        for ex_num, prob_key in [(1, "problem1"), (2, "problem2")]:
            counters[prefix] = counters.get(prefix, 0) + 1
            num = counters[prefix]
            ex_id  = f"{prefix}-A{num:03d}"
            ex_fn  = f"{ex_id}.png"
            problem = problems.get(prob_key, "")

            new_rows.append({"id": ex_id, "filename": ex_fn, "slug": slug,
                              "cluster": cluster, "keyword": keyword,
                              "example_num": ex_num, "problem": problem})
            print(f"  ✓ Ex{ex_num}: {ex_id} — {problem[:60]}...")

        done.add(slug)
        progress["done"] = list(done)
        progress["counters"] = counters
        save_progress(progress)
        time.sleep(0.3)

    all_rows = existing + new_rows
    if all_rows:
        save_excel(all_rows)

    print(f"\n✅ Done! New: {len(new_rows)} | Total: {len(all_rows)}")

if __name__ == "__main__":
    main()
