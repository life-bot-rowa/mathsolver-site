#!/usr/bin/env python3
"""Fix bad problems in article_examples.xlsx"""

import os, json, re, time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from openai import OpenAI
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openai", "openpyxl"], check=True)
    from openai import OpenAI

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "OPENAI_API_KEY")
MODEL = "gpt-4o-mini"
INPUT_FILE = Path("article_examples.xlsx")

# Keywords that indicate bad/irrelevant problems
BAD_KEYWORDS = [
    'geometry dash', 'game', 'player jumps', 'obstacle', 'level of',
    'online calculator', 'calculator app', 'math solver app', 'which tool',
    'ai tool', 'software', 'using an app', 'using a calculator',
    'navigate through', 'platform in geometry'
]

def is_bad_problem(problem, keyword):
    problem_lower = problem.lower()
    # Check if problem mentions game-related stuff
    if any(kw in problem_lower for kw in BAD_KEYWORDS):
        return True
    # Check if problem is about a tool/app rather than math
    if 'geometry dash' in keyword.lower() and 'geometry dash' in problem_lower:
        return True
    return False

def get_cluster_for_keyword(keyword):
    """Determine what kind of math problem to generate based on keyword context"""
    keyword_lower = keyword.lower()
    if 'geometry' in keyword_lower:
        return "geometry"
    if 'algebra' in keyword_lower:
        return "algebra"
    if 'calculus' in keyword_lower:
        return "calculus"
    if 'fraction' in keyword_lower or 'arithmetic' in keyword_lower:
        return "arithmetic and fractions"
    if 'statistic' in keyword_lower or 'probability' in keyword_lower:
        return "statistics"
    if 'equation' in keyword_lower:
        return "equation solving"
    if 'matrix' in keyword_lower or 'linear algebra' in keyword_lower:
        return "matrix and linear algebra"
    if 'trigonometry' in keyword_lower or 'trig' in keyword_lower:
        return "trigonometry"
    if 'physics' in keyword_lower:
        return "physics"
    return "general math"

def generate_replacement(client, keyword, cluster, example_num):
    math_topic = get_cluster_for_keyword(keyword)
    
    prompt = f"""Create 1 math problem for Example {example_num} about "{keyword}".
The problem must be a REAL MATH problem about {math_topic} — with actual numbers to calculate.
NOT about games, apps, or software.
Example {example_num}: {"easier" if example_num == 1 else "harder"} problem.
Plain text only, no LaTeX.

Return ONLY JSON: {{"problem": "the math problem"}}"""

    try:
        r = client.chat.completions.create(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.7,
            max_tokens=200,
        )
        raw = r.choices[0].message.content.strip()
        raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'\s*```\s*$', '', raw, flags=re.MULTILINE)
        return json.loads(raw).get("problem", "")
    except:
        return ""

def main():
    if OPENAI_API_KEY == "OPENAI_API_KEY":
        print("❌ Set OPENAI_API_KEY")
        return

    client = OpenAI(api_key=OPENAI_API_KEY)
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    fixed = 0
    for row in ws.iter_rows(min_row=2):
        screenshot_id = row[0].value
        keyword       = row[4].value or ""
        example_num   = row[5].value or 1
        problem       = row[6].value or ""

        if not screenshot_id or not problem:
            continue

        if is_bad_problem(problem, keyword):
            cluster = row[3].value or ""
            print(f"Fixing {screenshot_id} | {keyword}")
            print(f"  Old: {problem[:60]}...")

            new_problem = generate_replacement(client, keyword, cluster, example_num)
            if new_problem:
                row[6].value = new_problem
                print(f"  New: {new_problem[:60]}...")
                fixed += 1
            time.sleep(0.3)

    wb.save(INPUT_FILE)
    print(f"\n✅ Fixed {fixed} problems in {INPUT_FILE}")

if __name__ == "__main__":
    main()
